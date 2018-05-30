import numpy as np
import matplotlib as mpl
mpl.use('TkAgg')

import matplotlib.pyplot as plt
from sklearn.cluster import MiniBatchKMeans, KMeans
from sklearn.mixture import GaussianMixture
from sklearn.cluster import DBSCAN
from sklearn import metrics
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics.pairwise import pairwise_distances_argmin
from openpyxl import load_workbook
from sklearn.decomposition import IncrementalPCA
import csv
from functools import reduce
import statsmodels.api as sm
from statsmodels.formula.api import ols
import sys
import pandas as pd

# read the data from the excel sheet

# NASA TLX
wb = load_workbook('CompiledAllData_no_titles.xlsx', data_only=True)
worksheet = wb['Sheet1']
TLX_cols = ['F','G','H','I','J','K', 'L', 'M', 'N', 'O']
tlx_scores = []
for col in TLX_cols:
    tlx_scores.append([element.value for element in worksheet[col]])
tlx_scores =  np.transpose(np.array(tlx_scores, dtype=float))
filtered_tlx = tlx_scores[~np.isnan(tlx_scores).any(axis=1)]
tlx_attrs = filtered_tlx.shape[1]


EDA_data = []
with open('EDA_pca_short.csv', newline='') as csvfile:
     reader = csv.reader(csvfile, delimiter=',')
     for row in reader:
         EDA_data.append([float(i) for i in row])
eda = np.array(EDA_data, dtype=float)
print("EDA shape", eda.shape)
eda_attrs = eda.shape[1]

EEG_data = []
with open('EEG_pca_short.csv', newline='') as csvfile:
     reader = csv.reader(csvfile, delimiter=',')
     for row in reader:
         EEG_data.append([float(i) for i in row])
eeg = np.array(EEG_data, dtype=float)
print("EEG shape", eeg.shape)
eeg_attrs = eeg.shape[1]

HRV_data = []
with open('HRV_pca_short.csv', newline='') as csvfile:
     reader = csv.reader(csvfile, delimiter=',')
     for row in reader:
         HRV_data.append([float(i) for i in row])
hrv = np.array(HRV_data, dtype=float)
print("HRV shape", hrv.shape)
print("TLX shape", tlx_scores.shape)
hrv_attrs = hrv.shape[1]

eeg_hrv = np.concatenate((eeg, hrv), axis=1)
physical = np.concatenate((eeg_hrv, eda), axis=1)
all_data = np.concatenate((physical, filtered_tlx), axis=1)
datasets = [ filtered_tlx, eeg_hrv, physical, all_data]
titles = ["Nasa TLX", "Physiological (EEG+HRV)", "Physiological", "All"]

tlx_headers = ["TLX"+str(e) for e in range(tlx_attrs)]
eeg_headers = ["EEG"+str(e) for e in range(eeg_attrs)]
hrv_headers = ["HRV"+str(e) for e in range(hrv_attrs)]
eda_headers = ["EDA"+str(e) for e in range(eda_attrs)]
f = lambda a,b: a +","+ b
tlx_anova_header = reduce(f,tlx_headers)
eeg_hrv_anova_header = reduce(f,eeg_headers+hrv_headers)
physical_anova_header = reduce(f,eeg_headers+hrv_headers+eda_headers)
all_anova_header = reduce(f,eeg_headers+hrv_headers+eda_headers+tlx_headers)
anova_headers = [tlx_anova_header, eeg_hrv_anova_header, physical_anova_header, all_anova_header]
for i in range(len(datasets)):
    dataset_of_interest = datasets[i]
    filtered_data = dataset_of_interest[~np.isnan(dataset_of_interest).any(axis=1)]
    # print (filtered_data.shape, filtered_eda.shape)

    # Create the Models
    n_clusters = 3
    batch_size = 45
    k_means = KMeans(init='k-means++', n_clusters=n_clusters, n_init=10)
    mbk = MiniBatchKMeans(init='k-means++', n_clusters=n_clusters, batch_size=batch_size,
                          n_init=10, max_no_improvement=n_clusters, verbose=0)
    # gmm = GaussianMixture(n_components=3, covariance_type='full')
    # db = DBSCAN()

    # Fit the models
    # NASA TLX
    k_means.fit(filtered_data)
    mbk.fit(filtered_data)
    # gmm.fit(filtered_tlx)
    # db.fit(filtered_tlx)

    # Get the k means cluster labels
    k_means_cluster_centers = np.sort(k_means.cluster_centers_, axis=0)
    mbk_means_cluster_centers = np.sort(mbk.cluster_centers_, axis=0)
    k_means_labels = pairwise_distances_argmin(filtered_data, k_means_cluster_centers)
    mbk_means_labels = pairwise_distances_argmin(filtered_data, mbk_means_cluster_centers)
    order = pairwise_distances_argmin(k_means_cluster_centers,
                                      mbk_means_cluster_centers)
    # print(gmm.cluster_centers_)
    # print(db.cluster_centers_)

    # Plot the clusters
    # Load the labels
    # wb_p = load_workbook('CompiledTLXPerformance_PreviousBin.xlsx', data_only=True)
    wb_p = load_workbook('./CompiledTLXPerformance.xlsx', data_only=True)
    worksheet_p = wb_p['Sheet1']
    # performance_labels = [element.value for element in worksheet_p['AA']][1:]
    performance_labels = [element.value for element in worksheet_p['AA']]
    performance = [0  if v is None else v for v in performance_labels]
    performance = [1  if v=='Zero' else v for v in performance]
    # performance = [1  if v=='None' else v for v in performance]
    performance = [2  if v=='Low' else v for v in performance]
    performance = [3  if v=='High' else v for v in performance]

    # Apply PCA to the Data so we can plot it
    # Make this variable True to plot
    title = titles[i]
    plot = False
    if plot:
        ipca = IncrementalPCA(n_components=2, batch_size=3)
        plot_data = np.array(filtered_data, copy=True)
        print("plot data", plot_data.shape)
        ipca.fit(plot_data)
        ipca_plot_data = ipca.transform(plot_data)
        colors = ['#4EACC5', '#FF9C34', '#4E9A06', "#21EEFF"]
        markers = ['v', 'o', 'D']

        for index in range(len(ipca_plot_data)):
            plot_marker = markers[mbk_means_labels[index]]
            plot_color = colors[performance[index]]
            plt.plot(ipca_plot_data[index,0], ipca_plot_data[index,1], 'w',
                    markerfacecolor=plot_color, marker=plot_marker)
        plt.xlabel(title + ' PCA feature 1')
        plt.ylabel(title + ' PCA feature 2')
        plt.title('MBKMeans for '+title+' Data')
        plt.savefig('prevbins/MBKMeans_'+title+'.png')
        # plt.show()

    # Make the matrix for the cluster/performance improvement classes
    sample_size = filtered_data.shape[0]
    cross_matrix = [[' ', 'N/A', 'Low improvement', 'Med improvement', 'High improvement', 'Total'],
              ['Cluster 1', 0.0, 0.0, 0.0, 0.0, 0.0],
              ['Cluster 2', 0.0, 0.0, 0.0, 0.0, 0.0],
              ['Cluster 3', 0.0, 0.0, 0.0, 0.0, 0.0],
              ['Total', 0.0, 0.0, 0.0, 0.0, 1]]
    # Fill the matrix
    for index in range(sample_size):
        cross_matrix[mbk_means_labels[index]+1][performance[index]+1]+=1.0
    for index_1 in range(1,4):
        for index_2 in range(1,5):
            cross_matrix[index_1][index_2] /= sample_size
    # fill the sums
    cluster_sum = np.zeros(4)
    for index_1 in range(1,4):
        cross_matrix[index_1][5] = sum(cross_matrix[index_1][1:5])
        cluster_sum += np.array(cross_matrix[index_1][1:5])
    for index_2 in range(4):
        cross_matrix[4][index_2+1] = cluster_sum[index_2]
    # csv_file = open('prevbins/'+title+'.csv', 'w')
    csv_file = open(title+'.csv', 'w')
    with csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(cross_matrix)

    continue

    # Make the matrix for ANOVA analysis
    if False:
        anova_data =  np.concatenate((filtered_data,np.transpose([performance])),axis=1)
        anova_data =  np.concatenate((anova_data,np.transpose([mbk_means_labels])),axis=1)
        np.savetxt('prevbins/'+title+'_anova.csv', anova_data, delimiter=',',\
                header=anova_headers[i]+", Performace, Cluster", comments="")

    data = pd.read_csv('prevbins/'+title+'_anova.csv')
    # Do an anova for each one of the features in the cluster
    orig_stdout = sys.stdout
    print(title)
    f = open('prevbins/'+title+'_anova.txt', 'w')
    sys.stdout = f
    for elem in anova_headers[i].split(','):
        mod = ols(elem+' ~ Ctag', data=data).fit()
        # mod = sm.OLS(data[[elem]], data[['Cluster']]).fit()
        aov_table = sm.stats.anova_lm(mod, typ=2)
        print('%%%%%%%%%%%%%% ANOVA FOR FEATURE: ', elem)
        print(aov_table)
        esq_sm = aov_table['sum_sq'][0]/(aov_table['sum_sq'][0]+aov_table['sum_sq'][1])
        print('Effect Size:', esq_sm)
        print('/////////////////////////////////////////')
    sys.stdout = orig_stdout
    f.close()
